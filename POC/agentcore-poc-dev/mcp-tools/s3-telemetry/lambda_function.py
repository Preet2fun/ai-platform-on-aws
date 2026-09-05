"""MCP Tool: S3 Telemetry — Read anomaly telemetry data from S3 Excel files."""
import json
import io
import os
import boto3
import openpyxl


BUCKET = os.environ.get("RCA_S3_BUCKET", "")
PREFIX = "telemetry/"


def lambda_handler(event, context):
    # AgentCore Gateway: tool name in context, input as event directly
    delimiter = "___"
    try:
        original_tool_name = context.client_context.custom["bedrockAgentCoreToolName"]
        tool_name = original_tool_name[original_tool_name.index(delimiter) + len(delimiter):]
    except (AttributeError, KeyError, ValueError):
        tool_name = event.pop("toolName", event.pop("name", "read_telemetry"))
    tool_input = event

    handlers = {
        "read_telemetry": _read_telemetry,
        "list_telemetry_files": _list_files,
    }

    handler = handlers.get(tool_name)
    if not handler:
        return _response(f"Unknown tool: {tool_name}. Available: {list(handlers.keys())}")

    try:
        result = handler(tool_input)
        return _response(result)
    except Exception as e:
        return _response(f"Error: {str(e)}")


def _list_files(params):
    s3 = boto3.client("s3")
    resp = s3.list_objects_v2(Bucket=BUCKET, Prefix=PREFIX)
    files = [obj["Key"].split("/")[-1] for obj in resp.get("Contents", []) if obj["Key"].endswith(".xlsx")]
    return json.dumps({"files": files, "count": len(files), "bucket": BUCKET})


def _read_telemetry(params):
    s3 = boto3.client("s3")
    filename = params.get("filename", "")
    max_rows = params.get("max_rows", 50)

    # If no filename, read the first available
    if not filename:
        resp = s3.list_objects_v2(Bucket=BUCKET, Prefix=PREFIX)
        files = [obj["Key"] for obj in resp.get("Contents", []) if obj["Key"].endswith(".xlsx")]
        if not files:
            return json.dumps({"error": "No telemetry files found"})
        key = files[0]
    else:
        key = f"{PREFIX}{filename}" if not filename.startswith(PREFIX) else filename

    obj = s3.get_object(Bucket=BUCKET, Key=key)
    wb = openpyxl.load_workbook(io.BytesIO(obj["Body"].read()), read_only=True, data_only=True)

    results = []
    for sheet_name in wb.sheetnames[:2]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) for h in rows[0] if h]
        data_rows = rows[1:max_rows + 1]

        anomaly_count = sum(1 for r in rows[1:] for v in r if isinstance(v, str) and "anomaly" in v.lower())

        results.append({
            "file": key.split("/")[-1],
            "sheet": sheet_name,
            "headers": headers,
            "total_rows": len(rows) - 1,
            "sample_rows": [list(r) for r in data_rows[:5]],
            "anomaly_count": anomaly_count,
        })

    wb.close()
    return json.dumps({"telemetry": results}, default=str)


def _response(content):
    return {"content": [{"text": content if isinstance(content, str) else json.dumps(content, default=str)}]}
