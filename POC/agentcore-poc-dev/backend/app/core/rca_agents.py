"""RCA Sub-Agents — specialist agents for root cause analysis.

Each agent has its own optimised prompt + tools following AWS best practices:
IDENTITY → TASK → CONSTRAINTS → OUTPUT FORMAT
"""
import io
import os
import logging
from datetime import datetime, timedelta

import boto3
from langchain_aws import ChatBedrockConverse
from langchain_core.messages import HumanMessage
from langchain_core.tools import tool
from langgraph.prebuilt import create_react_agent

from app.core.config import settings

logger = logging.getLogger(__name__)

MODEL_ID = settings.MODEL
_session = None


def _boto():
    global _session
    if _session is None:
        _session = boto3.Session(region_name=settings.AWS_REGION)
    return _session


def _llm():
    return ChatBedrockConverse(model_id=MODEL_ID, region_name=settings.AWS_REGION, max_tokens=2048, temperature=0)


# ─── Tools ───────────────────────────────────────────────────────────────────

@tool
def read_telemetry(query: str) -> str:
    """Read telemetry data from S3 CSV/Excel files. Returns raw metric data with values, bounds, and anomaly flags."""
    try:
        import openpyxl
        s3 = _boto().client("s3", region_name=settings.AWS_REGION)
        bucket = os.environ.get("RCA_S3_BUCKET", "")
        resp = s3.list_objects_v2(Bucket=bucket, Prefix="telemetry/")
        files = [obj["Key"] for obj in resp.get("Contents", []) if obj["Key"].endswith(".xlsx")]
        results = []
        for key in files:
            obj = s3.get_object(Bucket=bucket, Key=key)
            wb = openpyxl.load_workbook(io.BytesIO(obj["Body"].read()), read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    continue
                headers = [str(h) for h in rows[0] if h]
                anomalies = sum(1 for r in rows[1:] for v in r if isinstance(v, str) and "anomaly" in v.lower())
                results.append(f"Source: {key.split('/')[-1]} | Sheet: {sheet}\nHeaders: {headers}\nTotal rows: {len(rows)-1} | Anomalies: {anomalies}\nSample: {rows[1:8]}")
            wb.close()
        return "\n\n".join(results) if results else "No telemetry data found."
    except Exception as e:
        return f"Error reading telemetry: {e}"


@tool
def query_cloudwatch(query: str) -> str:
    """Query CloudWatch for infrastructure metrics and alarms. Returns current alarm states and metric values."""
    try:
        cw = _boto().client("cloudwatch")
        resp = cw.describe_alarms(MaxRecords=15)
        alarms = resp.get("MetricAlarms", [])
        in_alarm = [a for a in alarms if a["StateValue"] == "ALARM"]
        lines = [f"Total alarms: {len(alarms)}, Active: {len(in_alarm)}"]
        for a in alarms[:15]:
            lines.append(f"  [{a['StateValue']}] {a['AlarmName']} — {a['Namespace']}/{a['MetricName']} {a['ComparisonOperator']} {a['Threshold']}")
        return "\n".join(lines)
    except Exception as e:
        return f"CloudWatch error: {e}"


@tool
def lookup_cloudtrail(query: str) -> str:
    """Look up recent CloudTrail events (last 24h). Returns API calls, deployments, and config changes."""
    try:
        ct = _boto().client("cloudtrail", region_name=settings.AWS_REGION)
        resp = ct.lookup_events(StartTime=datetime.utcnow() - timedelta(hours=24), EndTime=datetime.utcnow(), MaxResults=15)
        events = resp.get("Events", [])
        lines = [f"CloudTrail events (24h): {len(events)}"]
        for e in events[:15]:
            lines.append(f"  [{e.get('EventTime','')}] {e.get('EventName','')} by {e.get('Username','')} — {e.get('EventSource','')}")
        return "\n".join(lines) if events else "No CloudTrail events found in last 24 hours."
    except Exception as e:
        return f"CloudTrail error: {e}"


# ─── Agent Prompts ───────────────────────────────────────────────────────────

INCIDENT_DETECTION_PROMPT = """## IDENTITY
You are an Incident Detection Agent for an enterprise IT observability platform.

## TASK
Given an alert or telemetry data, validate whether this is a real anomaly:
1. Compare values against provided bounds/baselines
2. Determine anomaly confidence
3. Classify incident type and severity

## CONSTRAINTS
- MUST only flag values that numerically breach provided bounds
- MUST NOT speculate about causes — detection only
- MUST cite specific values and timestamps

## OUTPUT FORMAT
Incident: [Confirmed/Not Confirmed]
Type: [Memory Exhaustion | CPU Spike | Latency | Error Rate | Availability]
Severity: [Critical | High | Medium | Low]
Confidence: [0-100]%
Evidence: [specific values, timestamps, deviation percentage]"""

TELEMETRY_ANALYSIS_PROMPT = """## IDENTITY
You are a Telemetry Analysis Agent for an enterprise IT observability platform.

## TASK
Given metric data with values and bounds:
1. Identify all breach points (value > upper_bound or value < lower_bound)
2. Classify pattern: Isolated (single), Sustained (3+ consecutive), Progressive (worsening), Recurring (periodic)
3. Quantify: peak magnitude, duration, frequency

## CONSTRAINTS
- MUST reference specific timestamps and numeric values
- MUST only classify as anomalous if bounds are numerically breached
- MUST NOT confuse correlation with causation

## OUTPUT FORMAT
Anomalies Found: [count]
Pattern: [Isolated | Sustained | Progressive | Recurring]
Peak: [value] at [timestamp] (bound: [threshold], breach: [magnitude])
Duration: [time window]
Trend: [Worsening | Stable | Recovering]"""

INFRASTRUCTURE_PROMPT = """## IDENTITY
You are an Infrastructure Analysis Agent for an enterprise IT observability platform.

## TASK
Given infrastructure metrics:
1. Check CPU, Memory, Disk, Network against normal ranges
2. Determine if infrastructure IS or IS NOT contributing to the incident
3. Explicitly confirm or eliminate infrastructure as cause

## CONSTRAINTS
- MUST check all dimensions available in the data
- MUST explicitly state "eliminated" or "contributing" for each
- MUST NOT assume — only report what metrics show

## OUTPUT FORMAT
CPU: [Normal | Elevated | Critical] — [value]
Memory: [Normal | Elevated | Critical] — [value]
Disk: [Normal | Elevated | Critical] — [value]
Network: [Normal | Elevated | Critical] — [value]
Conclusion: Infrastructure [is | is not] contributing to the incident"""

CHANGE_ANALYSIS_PROMPT = """## IDENTITY
You are a Change Analysis Agent for an enterprise IT observability platform.

## TASK
Given event history:
1. List changes in the 24 hours before the incident
2. For each: who, what, when
3. Assess correlation strength based on timing

## CONSTRAINTS
- MUST focus on changes BEFORE the incident
- MUST NOT assume causation from correlation alone
- If no changes found, state "No changes detected"

## OUTPUT FORMAT
Changes Found: [count] in 24h window
[timestamp] — [event] by [user] — Correlation: [Strong | Weak | None]
Assessment: [changes are | are not] likely related to the incident"""

RCA_ENGINE_PROMPT = """## IDENTITY
You are the RCA Engine. You produce the final incident report from investigation findings.

## TASK
Synthesize all findings into a structured, actionable report.

## CONSTRAINTS
- MUST back every claim with evidence from the investigation
- MUST NOT speculate beyond the data
- If confidence < 70%, MUST state what additional data would help
- MUST include both immediate fix and preventive measure

## OUTPUT FORMAT
**Incident Summary**: [one line]
**Root Cause**: [originating event with evidence]
**Confidence**: [%] — [reasoning]
**Causal Chain**: [A] → [B] → [C]
**Affected**: [service/host]
**Severity**: [Critical | High | Medium | Low]
**Remediation**:
  - Immediate: [fix now]
  - Preventive: [stop recurrence]
**Resolution ETA**: [estimate]"""


# ─── Build Sub-Agents ────────────────────────────────────────────────────────

_agents = {}


def _get_agent(name):
    if name in _agents:
        return _agents[name]

    llm = _llm()
    configs = {
        "incident_detection": (INCIDENT_DETECTION_PROMPT, [read_telemetry]),
        "telemetry_analysis": (TELEMETRY_ANALYSIS_PROMPT, [read_telemetry]),
        "infrastructure": (INFRASTRUCTURE_PROMPT, [query_cloudwatch]),
        "change_analysis": (CHANGE_ANALYSIS_PROMPT, [lookup_cloudtrail]),
        "rca_engine": (RCA_ENGINE_PROMPT, []),
    }

    prompt, tools = configs[name]
    agent = create_react_agent(llm, tools, prompt=prompt)
    _agents[name] = agent
    return agent


async def run_agent(name: str, input_text: str) -> str:
    """Run a named RCA sub-agent and return its output."""
    agent = _get_agent(name)
    result = await agent.ainvoke({"messages": [HumanMessage(content=input_text)]})
    return result["messages"][-1].content
